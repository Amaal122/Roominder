import os
from openpyxl import Workbook
from sqlalchemy.orm import Session
try:
    from .auth import SeekerProfile
except ImportError:
    SeekerProfile = None

def export_seeker_profile_to_excel(db: Session, seeker_profile):
    from openpyxl import load_workbook
    # Use a single file for all profiles
    filename = "seeker_profiles.xlsx"
    backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    filepath = os.path.join(backend_dir, filename)

    headers = [
        "id", "user_id", "looking_for", "location", "radius", "age", "gender", "occupation", "image_url",
        "sleep_schedule", "cleanliness", "social_life", "guests", "work_style"
    ]
    data_row = [
        seeker_profile.id,
        seeker_profile.user_id,
        seeker_profile.looking_for,
        seeker_profile.location,
        seeker_profile.radius,
        seeker_profile.age,
        seeker_profile.gender,
        seeker_profile.occupation,
        seeker_profile.image_url,
        seeker_profile.sleep_schedule,
        seeker_profile.cleanliness,
        seeker_profile.social_life,
        seeker_profile.guests,
        seeker_profile.work_style,
    ]

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        # Only add headers if the file is empty
        if ws.max_row == 0:
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "SeekerProfile"
        ws.append(headers)

    ws.append(data_row)

    try:
        wb.save(filepath)
    except Exception as e:
        print(f"Failed to save Excel file: {filepath} - {e}")
    return filepath
